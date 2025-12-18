import pandas as pd
import numpy as np
from sklearn.metrics import mean_squared_error, mean_absolute_error
from rectools.metrics import (
    Precision, Recall, F1Beta, MAP, NDCG, MRR, HitRate,
    IntraListDiversity, CatalogCoverage, calc_metrics
)

from DataHandler2 import load_and_process_data
import os
import tempfile
from datetime import datetime
import warnings
from sklearn.exceptions import DataConversionWarning
import contextlib
from io import StringIO
import sys
from openpyxl import load_workbook
import re

warnings.filterwarnings("ignore", category=DataConversionWarning, module="sklearn.metrics.pairwise")

# Import from your custom modules
from Plotting import plot_individual_metric_charts, plot_rating_distribution, plot_metrics_vs_k_from_directory
from Diagnostics import _print_data_diagnostics


class Tee:
    """Capture terminal output while still displaying it"""

    def __init__(self, terminal, buffer):
        self.terminal = terminal
        self.buffer = buffer

    def write(self, message):
        self.terminal.write(message)
        self.buffer.write(message)

    def flush(self):
        self.terminal.flush()
        self.buffer.flush()


def sanitize_for_excel(text):
    """Sanitize text to prevent Excel errors"""
    if not text:
        return ""
    text = re.sub(r'[\x00-\x08\x0B-\x1F\x7F]', '', text)
    if text.startswith(('=', '+', '-', '@')):
        text = "'" + text
    if len(text) > 32700:
        text = text[:32700] + " [TRUNCATED]"
    return text


def validate_files(config_dict):
    """Validate all required files before processing"""
    errors = []
    print("\n🔍 Validating all input files...")

    # Check catalog
    if not os.path.exists(config_dict['CATALOG_PATH']):
        errors.append(f"❌ Catalog file not found: {config_dict['CATALOG_PATH']}")
    else:
        try:
            pd.read_csv(config_dict['CATALOG_PATH'], nrows=1)
            print(f"✅ Catalog: {config_dict['CATALOG_PATH']}")
        except Exception as e:
            errors.append(f"❌ Cannot load catalog: {config_dict['CATALOG_PATH']} - {e}")

    # Check ground truth
    if not os.path.exists(config_dict['GROUND_TRUTH']):
        errors.append(f"❌ Ground truth file not found: {config_dict['GROUND_TRUTH']}")
    else:
        try:
            pd.read_csv(config_dict['GROUND_TRUTH'], encoding='latin1', nrows=1)
            print(f"✅ Ground truth: {config_dict['GROUND_TRUTH']}")
        except Exception as e:
            errors.append(f"❌ Cannot load ground truth: {config_dict['GROUND_TRUTH']} - {e}")

    # Check item features if ILD enabled
    if config_dict.get('CALCULATE_ILD', False):
        item_features_path = config_dict.get('ITEM_FEATURES_PATH')
        if item_features_path and os.path.exists(item_features_path):
            try:
                pd.read_csv(item_features_path, engine='python', on_bad_lines='skip', nrows=1)
                print(f"✅ Item features: {item_features_path}")
            except Exception as e:
                errors.append(f"❌ Cannot load item features: {item_features_path} - {e}")
        elif item_features_path:
            errors.append(f"❌ Item features file not found: {item_features_path}")

    # Check model prediction files
    for predictions_path, source_name in config_dict.get('MODELS', []):
        if not os.path.exists(predictions_path):
            errors.append(f"❌ Model file not found for '{source_name}': {predictions_path}")
        else:
            try:
                dataset_type = config_dict.get('dataset_type', 'movies')
                if dataset_type == "books":
                    pd.read_csv(predictions_path, encoding='latin1', nrows=1)
                else:
                    valid_lines = []
                    with open(predictions_path, 'r', encoding='latin1') as f:
                        header_line = None
                        for line_num, line in enumerate(f):
                            stripped = line.strip()
                            if line_num == 0 and (stripped.startswith('userId') or stripped.startswith('user_id')):
                                header_line = line
                            if stripped != '' and not stripped.startswith('#'):
                                valid_lines.append(line)
                            if line_num > 100:
                                break

                    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='latin1') as temp:
                        if header_line:
                            temp.write(header_line)
                        temp.writelines(valid_lines[1:] if header_line else valid_lines)
                        temp_path = temp.name

                    try:
                        pd.read_csv(temp_path, nrows=1)
                    finally:
                        os.unlink(temp_path)

                print(f"✅ Model '{source_name}': {predictions_path}")
            except Exception as e:
                errors.append(f"❌ Cannot load model file for '{source_name}': {predictions_path} - {e}")

    if errors:
        print("\n" + "=" * 60)
        print(" FILE VALIDATION ERRORS FOUND - STOPPING EXECUTION ")
        print("=" * 60)
        for error in errors:
            print(error)
        print("\nPlease fix the issues above and run again.")
        raise SystemExit(1)

    print("\n✅ All files validated successfully!\n")


def calculate_all_metrics(catalog, data_handler, threshold=4.0, k=5, item_features=None,
                          model_name="Unknown", calculate_ild=True):
    results = {}

    print(f"Calculating RMSE and MAE for {model_name}")
    rmse, mae = _calculate_rating_metrics(data_handler, model_name)
    results["RMSE"] = rmse
    results["MAE"] = mae

    print(f"Calculating top-{k} metrics for {model_name}")

    relevant_interactions = data_handler.interactions[
        data_handler.interactions['weight'] >= threshold
        ].copy()

    gt_items = data_handler.interactions["item_id"].unique()
    pred_items = data_handler.recommendations["item_id"].unique()
    catalog = np.union1d(gt_items, pred_items)
    catalog_size = len(catalog)

    metrics = {
        f'Precision@{k}': Precision(k=k),
        f'Recall@{k}': Recall(k=k),
        f'F1Beta@{k}': F1Beta(k=k, beta=1.0),
        f'MAP@{k}': MAP(k=k),
        f'NDCG@{k}': NDCG(k=k),
        f'MRR@{k}': MRR(k=k),
        f'CatalogCoverage@{k}': CatalogCoverage(k=k),
        f'HitRate@{k}': HitRate(k=k),
    }

    try:
        metrics_values = calc_metrics(
            metrics=metrics,
            reco=data_handler.recommendations,
            interactions=relevant_interactions,
            catalog=catalog,
            prev_interactions=None,
        )

        results[f"HitRate@{k}"] = metrics_values[f'HitRate@{k}']
        results[f"Precision@{k}"] = metrics_values[f'Precision@{k}']
        results[f"Recall@{k}"] = metrics_values[f'Recall@{k}']
        results[f"F1@{k}"] = metrics_values[f'F1Beta@{k}']
        results[f"MAP@{k}"] = metrics_values[f'MAP@{k}']
        results[f"NDCG@{k}"] = metrics_values[f'NDCG@{k}']
        results[f"MRR@{k}"] = metrics_values[f'MRR@{k}']
        results[f"Coverage@{k}"] = metrics_values[f'CatalogCoverage@{k}'] / catalog_size

    except Exception as e:
        print(f"Warning: Error calculating ranking metrics for {model_name}: {e}")
        print("Filling with NaN values and continuing")
        results[f"HitRate@{k}"] = np.nan
        results[f"Precision@{k}"] = np.nan
        results[f"Recall@{k}"] = np.nan
        results[f"F1@{k}"] = np.nan
        results[f"MAP@{k}"] = np.nan
        results[f"NDCG@{k}"] = np.nan
        results[f"MRR@{k}"] = np.nan
        results[f"Coverage@{k}"] = np.nan

    if calculate_ild and item_features is not None and not item_features.empty:
        print(f"Calculating ILD@{k} for {model_name}")
        results[f"ILD@{k}_Jaccard"] = _calculate_ild(data_handler, item_features, k, metric='jaccard')
        results[f"ILD@{k}_Cosine"] = _calculate_ild(data_handler, item_features, k, metric='cosine')
    else:
        results[f"ILD@{k}_Jaccard"] = np.nan
        results[f"ILD@{k}_Cosine"] = np.nan

    top_k_recos = data_handler.recommendations[data_handler.recommendations['rank'] <= k]

    print(f"Calculating Reverse Gini for {model_name}")
    results['Reverse Gini'] = _calculate_reverse_gini(top_k_recos)

    return results


def _calculate_rating_metrics(data_handler, model_name="Unknown"):
    try:
        merged = pd.merge(
            data_handler.predictions,
            data_handler.interactions,
            on=['user_id', 'item_id'],
            suffixes=('_pred', '_gt')
        )

        merged_clean = merged[['weight_gt', 'weight_pred']].dropna()

        if merged_clean.empty:
            print(f"Warning for {model_name}: After merge, no valid rating pairs found (all NaN).")
            return np.nan, np.nan

        rmse = np.sqrt(mean_squared_error(
            merged_clean['weight_gt'],
            merged_clean['weight_pred']
        ))
        mae = mean_absolute_error(
            merged_clean['weight_gt'],
            merged_clean['weight_pred']
        )
        return rmse, mae

    except Exception as e:
        print(f"\nError in RMSE/MAE calculation for {model_name}: {e}")
        print("   Returning NaN for RMSE and MAE to allow other metrics to continue.\n")
        return np.nan, mae


def _calculate_ild(data_handler, item_features, k, metric="cosine"):
    """
    Simplified and corrected ILD calculation.
    Higher values = more diversity (0-1 range).
    """
    try:
        recos = data_handler.recommendations.copy()
        recos['item_id'] = recos['item_id'].astype(str)
        recos = recos[recos['rank'] <= k]

        # Filter to items with features
        available_items = list(set(recos['item_id'].unique()) & set(item_features.index.astype(str)))
        if not available_items:
            print("⚠️  No items have features!")
            return np.nan

        # Create distance matrix
        feature_matrix = item_features.loc[available_items].values

        if metric == 'cosine':
            norms = np.linalg.norm(feature_matrix, axis=1, keepdims=True)
            norms[norms == 0] = 1
            feature_normalized = feature_matrix / norms
            similarity = np.dot(feature_normalized, feature_normalized.T)
            distance_matrix = 1 - np.clip(similarity, 0, 1)
            np.fill_diagonal(distance_matrix, 0)

        elif metric == 'jaccard':
            intersection = feature_matrix @ feature_matrix.T
            sum_features = feature_matrix.sum(axis=1)
            union = sum_features[:, np.newaxis] + sum_features[np.newaxis, :] - intersection
            jaccard_similarity = intersection / np.maximum(union, 1)
            distance_matrix = 1 - jaccard_similarity
            np.fill_diagonal(distance_matrix, 0)

        else:
            raise ValueError(f"Unsupported metric: {metric}")

        item_to_idx = {item: idx for idx, item in enumerate(available_items)}

        # Calculate ILD per user
        ild_values = []

        for user_id, user_recos in recos.groupby('user_id'):
            user_items = user_recos['item_id'].astype(str).tolist()
            user_items = [item for item in user_items if item in item_to_idx]

            if len(user_items) >= 2:
                idxs = [item_to_idx[item] for item in user_items]
                user_distances = distance_matrix[np.ix_(idxs, idxs)]
                upper_tri = user_distances[np.triu_indices_from(user_distances, k=1)]

                if len(upper_tri) > 0:
                    user_ild = upper_tri.mean()
                    ild_values.append(user_ild)

        final_ild = np.mean(ild_values) if ild_values else np.nan

        if final_ild < 0.5:
            print(f"WARNING: ILD={final_ild:.3f} is suspiciously low for {metric}")

        return final_ild

    except Exception as e:
        print(f"ILD calculation failed for {metric}: {e}")
        return np.nan


def _calculate_reverse_gini(recommendations):
    item_counts = recommendations['item_id'].value_counts()
    n = len(item_counts)

    if n < 2:
        print("n < 2, returning 1.0")
        return 1.0

    sorted_counts = np.sort(item_counts.values)
    total = sorted_counts.sum()
    proportions = sorted_counts / total

    i = np.arange(1, n + 1)
    weighted_sum = np.sum(i * proportions)
    gini = (2 * weighted_sum - (n + 1)) / (n - 1)

    gini = np.clip(gini, 0.0, 1.0)
    result = 1 - gini
    return result


def display_metrics_table(metrics_dict, source_name="Model", k=5):
    overall_metrics = ["RMSE", "MAE", "Reverse Gini"]
    topk_metrics = [
        f"HitRate@{k}",
        f"Precision@{k}",
        f"Recall@{k}",
        f"F1@{k}",
        f"NDCG@{k}",
        f"MAP@{k}",
        f"MRR@{k}",
        f"Coverage@{k}",
        f"ILD@{k}_Jaccard",
        f"ILD@{k}_Cosine"
    ]
    metric_order = overall_metrics + topk_metrics

    ordered_results = {metric: metrics_dict.get(metric, np.nan) for metric in metric_order}
    df = pd.DataFrame([ordered_results], index=[source_name])
    df_display = df.round(4)

    print(df_display.to_string())
    return df


def save_metrics_table_as_file(metrics_df, filename="metrics_results"):
    metrics_df.to_excel(f"{filename}.xlsx")
    print(f"Saved: {filename}.xlsx")


def load_item_features(items_path, dataset_type="movies"):
    print(f"Loading item features for {dataset_type}")

    items = pd.read_csv(items_path, engine='python', on_bad_lines='skip')

    if 'itemId' not in items.columns and 'item_id' in items.columns:
        items = items.rename(columns={'item_id': 'itemId'})

    items['itemId'] = items['itemId'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    items = items.drop_duplicates(subset=['itemId'], keep='first')
    print(f"After deduplication: {len(items)} unique items")

    items['genres'] = items['genres'].fillna('Unknown')
    items['genres_list'] = items['genres'].str.split('|')

    all_genres = set()
    for genres in items['genres_list']:
        if isinstance(genres, list):
            for g in genres:
                clean_genre = g.strip() if isinstance(g, str) else ''
                if clean_genre and clean_genre.lower() != 'unknown':
                    all_genres.add(clean_genre)

    all_genres = sorted(list(all_genres))
    print(f"Found {len(all_genres)} genres: {all_genres[:10]}...")

    unique_ids = items['itemId'].unique()
    item_features = pd.DataFrame(0, index=unique_ids, columns=all_genres)
    item_features.index.name = 'item_id'

    items_with_features = 0
    for idx, row in items.iterrows():
        if isinstance(row['genres_list'], list):
            valid_genres = [g.strip() for g in row['genres_list']
                            if g and g.strip() in all_genres]
            if valid_genres:
                for genre in valid_genres:
                    item_features.at[row['itemId'], genre] = 1
                items_with_features += 1

    print(f"Processed features for {items_with_features}/{len(items)} items")
    print(f"Matrix shape: {item_features.shape}")
    print(f"Index unique: {item_features.index.is_unique}")
    return item_features


def run_model_comparison(ground_truth_path, sources, catalog, threshold=4.0, k=5,
                         item_features=None, output_prefix="comparison",
                         calculate_ild=True, dataset_type="movies"):
    all_results_df = pd.DataFrame()

    print("Metrics calculations")

    for predictions_path, source_name in sources:
        print(f"\nProcessing '{source_name}'")

        try:
            data = load_and_process_data(ground_truth_path, predictions_path, dataset_type=dataset_type, verbose=False)
        except Exception as e:
            print(f"Error loading data for {source_name}: {e}")
            print("Skipping this model")
            continue

        metrics = calculate_all_metrics(catalog, data, threshold, k, item_features,
                                        source_name, calculate_ild)
        source_df = display_metrics_table(metrics, source_name, k)
        all_results_df = pd.concat([all_results_df, source_df])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{output_prefix}_results_{timestamp}"
    save_metrics_table_as_file(all_results_df, filename)

    print(f"\nProcessed {len(all_results_df)} model(s) with k={k}")
    return all_results_df, filename


# ----- Main Execution Block -----
if __name__ == "__main__":
    # Import configuration
    from DataPaths import (
        THRESHOLD, K, CALCULATE_ILD,
        CATALOG_PATH, CATALOG,
        GROUND_TRUTH,
        MODELS,
        ITEM_FEATURES_PATH
    )

    # Load item features ONCE before the loop
    if CALCULATE_ILD:
        print("Loading item features for ILD calculation")
        ITEM_FEATURES = load_item_features(ITEM_FEATURES_PATH, dataset_type="books")
    else:
        print("Skipping item feature loading (ILD disabled)")
        ITEM_FEATURES = None

    # Process K values x to y
    for current_k in range(10, 9, -1):
        print(f"\n{'=' * 60}")
        print(f"Processing with k={current_k}")
        print(f"{'=' * 60}\n")

        # Capture output for this iteration
        stdout_buffer = StringIO()
        tee_stdout = Tee(sys.stdout, stdout_buffer)

        try:
            with contextlib.redirect_stdout(tee_stdout):
                # Validate files
                config_for_validation = {
                    'CATALOG_PATH': CATALOG_PATH,
                    'GROUND_TRUTH': GROUND_TRUTH,
                    'CALCULATE_ILD': CALCULATE_ILD,
                    'ITEM_FEATURES_PATH': ITEM_FEATURES_PATH,
                    'MODELS': MODELS,
                    'dataset_type': "books"
                }
                validate_files(config_for_validation)

                # Run diagnostics
                _print_data_diagnostics(GROUND_TRUTH, file_label="Ground Truth",
                                        threshold=THRESHOLD, is_ground_truth=True)

                # For predictions (add ground_truth_path parameter)
                for predictions_path, source_name in MODELS:
                    _print_data_diagnostics(predictions_path,
                                            file_label=f"Model '{source_name}'",
                                            threshold=THRESHOLD,
                                            is_ground_truth=False,
                                            ground_truth_path=GROUND_TRUTH)

                # Run comparison
                results, filename = run_model_comparison(
                    ground_truth_path=GROUND_TRUTH,
                    sources=MODELS,
                    threshold=THRESHOLD,
                    k=current_k,
                    item_features=ITEM_FEATURES,
                    output_prefix=f"kasia, 100k movies top{current_k}_comparison",
                    calculate_ild=CALCULATE_ILD,
                    catalog=CATALOG,
                    dataset_type="books"
                    #dataset_type = "movies"
                )

            # Save terminal output to Excel
            terminal_output = stdout_buffer.getvalue()
            excel_file = f"{filename}.xlsx"
            wb = load_workbook(excel_file)

            if "Terminal Output" in wb.sheetnames:
                wb.remove(wb["Terminal Output"])

            ws = wb.create_sheet("Terminal Output")
            for i, line in enumerate(terminal_output.split('\n'), 1):
                ws.cell(row=i, column=1, value=sanitize_for_excel(line))

            ws.column_dimensions['A'].width = 100
            ws.freeze_panes = 'A2'
            wb.save(excel_file)
            print(f"✅ Terminal output saved to {excel_file}")

        except Exception as e:
            print(f"\n❌ Error processing k={current_k}: {e}")
            print("   Continuing with next K value...")
            continue

    print(f"\n{'=' * 60}")
    print("All K values processed!")
    print(f"{'=' * 60}")